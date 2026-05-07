#!/usr/bin/env python3
"""
LinkedIn Job Scraper
Scrapes job postings from LinkedIn and saves to Excel.

Usage:
    python linkedin_scraper.py [--url URL] [--pages N] [--output FILE] [--wait N]

Requirements:
    pip install selenium openpyxl webdriver-manager
    Chrome browser must be installed.
"""

import argparse
import json
import random
import re
import sys
import time
from collections import Counter

# Force UTF-8 output on Windows so emoji/Turkish chars don't crash the terminal
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.common.exceptions import (
        TimeoutException,
        NoSuchElementException,
        StaleElementReferenceException,
        ElementClickInterceptedException,
    )
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install selenium openpyxl webdriver-manager")
    sys.exit(1)


DEFAULT_KEYWORDS = [
    "Yazılım Mühendisi",
    "Backend Developer",
    "Frontend Developer",
    "Full Stack Developer",
    "Data Scientist",
    "DevOps Engineer",
    "Machine Learning Engineer",
    "Siber Güvenlik Uzmanı",
    "Cloud Engineer",
    "Veri Bilimi",
    "QA Engineer",
    "Mobile Developer",
]

# ── ESCO broader-concept → LinkedIn job-title mapping ────────────────────────

_DIGITAL_CONCEPT_MAP = {
    "computer programming": "Software Developer",
    "software and applications development and analysis": "Software Engineer",
    "database and network design and administration": "Database Administrator",
    "ict project management": "IT Project Manager",
    "penetration testing": "Cybersecurity Engineer",
    "protecting ict devices": "Cybersecurity Analyst",
    "information security": "Information Security Analyst",
    "managing, gathering and storing digital data": "Data Engineer",
    "designing ict systems or applications": "Software Architect",
    "designing electrical or electronic systems": "Electronics Engineer",
    "electronics and automation": "Automation Engineer",
    "network engineering": "Network Engineer",
    "computer technology": "IT Specialist",
    "artificial intelligence": "AI Engineer",
    "machine learning": "Machine Learning Engineer",
    "data science": "Data Scientist",
    "cloud": "Cloud Engineer",
    "devops": "DevOps Engineer",
    "web development": "Web Developer",
    "mobile": "Mobile Developer",
    "robotics": "Robotics Engineer",
    "internet of things": "IoT Engineer",
    "digital game": "Game Developer",
    "computer use": "IT Support Specialist",
    "setting up computer systems": "Systems Administrator",
    "working with computers": "IT Specialist",
    "statistical analysis": "Data Analyst",
    "business intelligence": "Business Intelligence Analyst",
}

_GREEN_CONCEPT_MAP = {
    "renewable energy": "Renewable Energy Engineer",
    "electricity and energy": "Energy Engineer",
    "environmental sciences": "Environmental Engineer",
    "environmental policy": "Environmental Consultant",
    "environmental impact": "Environmental Impact Assessor",
    "sustainability": "Sustainability Specialist",
    "waste management": "Waste Management Specialist",
    "climate": "Climate Change Analyst",
    "energy efficiency": "Energy Efficiency Consultant",
    "circular economy": "Circular Economy Specialist",
    "water": "Water Resource Engineer",
    "carbon": "Carbon Management Specialist",
    "esg": "ESG Analyst",
    "green building": "Green Building Consultant",
    "biomass": "Biomass Energy Engineer",
    "solar": "Solar Energy Engineer",
    "wind power": "Wind Energy Engineer",
    "heat pump": "HVAC Engineer",
    "crop and livestock": "Agricultural Engineer",
    "complying with environmental": "Environmental Compliance Officer",
    "community sanitation": "Environmental Health Officer",
}

# Technology names to skip (too niche or not useful as LinkedIn job searches)
_SKIP_TECH_NAMES = {
    "haskell", "erlang", "kdevplatform", "kdevelop", "cryengine",
    "absorb", "maltego", "sas language", "kdevplatform",
}

def _is_searchable_tech_name(label: str) -> bool:
    """Return True if label is a short, recognisable technology/tool name.

    Rules:
    - Must start with a capital letter (proper noun / product name, e.g. Python, AWS).
    - At most 2 words (e.g. "Machine Learning" ok, "incremental development model" not).
    - Not in the niche-tool skip list.
    """
    if not label or len(label) > 30:
        return False
    if label.lower() in _SKIP_TECH_NAMES:
        return False
    words = label.split()
    if len(words) > 2:
        return False
    # Must begin with an uppercase letter — filters out generic phrases like
    # "network engineering", "incremental development", "information structure"
    return label[0].isupper()


def load_esco_keywords(
    digital_csv: str = "data/esco/digitalSkillsCollection_en.csv",
    green_csv: str = "data/esco/greenSkillsCollection_en.csv",
) -> list[str]:
    """
    Build LinkedIn search keywords from ESCO digital and green skill collections.

    Sources:
      1. Unique broaderConceptPT domain labels → mapped to job-title search terms.
      2. knowledge-type digital skill labels → technology/tool names used directly.

    Falls back to DEFAULT_KEYWORDS if CSV files are not found.
    """
    import csv

    keywords: list[str] = []
    seen: set[str] = set()

    def add(kw: str) -> None:
        if kw and kw.lower() not in seen:
            seen.add(kw.lower())
            keywords.append(kw)

    def map_concepts(broader_str: str, concept_map: dict) -> None:
        for concept in broader_str.split(" | "):
            c = concept.strip().lower()
            for key, job_title in concept_map.items():
                if key in c:
                    add(job_title)
                    break

    # Prefer using shared loader so occupations can contribute keywords if desired
    from src.utils.esco_loader import load_esco_taxonomy

    digital_path = Path(digital_csv)
    green_path = Path(green_csv)

    if not digital_path.exists() and not green_path.exists():
        print("  [ESCO] CSV files not found — using DEFAULT_KEYWORDS.")
        return DEFAULT_KEYWORDS

    # ── Digital skills ──────────────────────────────────────────────────────
    if digital_path.exists():
        # Use loader to ensure consistent parsing and to surface occupations if useful
        digital_uris, green_uris, uri_to_label, occupation_uri_to_label, occupation_to_skill_uris = load_esco_taxonomy(
            data_dir=Path("data/esco"),
            digital_csv=digital_path,
            green_csv=green_path,
            skills_csv=Path("data/esco/skills_en.csv"),
        )
        # iterate original CSV to preserve altLabels logic for building keywords
        with open(digital_path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                label = (row.get("preferredLabel") or "").strip()
                skill_type = (row.get("skillType") or "").strip()
                broader = (row.get("broaderConceptPT") or "").strip()

                if skill_type == "knowledge" and _is_searchable_tech_name(label):
                    add(label)

                map_concepts(broader, _DIGITAL_CONCEPT_MAP)

    # ── Green skills ─────────────────────────────────────────────────────────
    if green_path.exists():
        with open(green_path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                broader = (row.get("broaderConceptPT") or "").strip()
                map_concepts(broader, _GREEN_CONCEPT_MAP)

    if not keywords:
        print("  [ESCO] No keywords extracted — using DEFAULT_KEYWORDS.")
        return DEFAULT_KEYWORDS

    print(f"  [ESCO] Loaded {len(keywords)} search keywords from ESCO skill collections.")
    return keywords


def build_keyword_url(keyword: str) -> str:
    return (
        f"https://www.linkedin.com/jobs/search/?keywords={quote(keyword)}"
        "&location=T%C3%BCrkiye&f_TPR=r604800&position=1&pageNum=0"
    )


def _normalize_job_url(url: str) -> str:
    if not url:
        return ""
    m = re.search(r"/jobs/view/(\d+)", url)
    if m:
        return f"linkedin_job_{m.group(1)}"
    return url.split("?")[0]


def load_existing_urls(paths: list[str]) -> set[str]:
    """Load normalised job URLs from existing Excel files for deduplication."""
    seen: set[str] = set()
    for p in paths:
        path = Path(p)
        if not path.exists():
            continue
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(path, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = [c.value for c in next(ws.iter_rows(max_row=1))]
                url_col = next(
                    (i for i, h in enumerate(headers) if h and h.lower() == "url"), None
                )
                if url_col is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    url = row[url_col]
                    if url and isinstance(url, str) and "/jobs/view/" in url:
                        seen.add(_normalize_job_url(url))
            wb.close()
        except Exception as e:
            print(f"  Warning: could not read {path.name} for dedup: {e}")
    return seen


SELECTORS = {
    "job_cards": [
        "li.jobs-search-results__list-item",
        "div.job-search-card",
        "li[data-occludable-job-id]",
        "li.scaffold-layout__list-item",
    ],
    "job_title": [
        "h1.job-details-jobs-unified-top-card__job-title",
        "h1.t-24",
        "h2.job-details-jobs-unified-top-card__job-title",
        "h1[data-test-id='job-title']",
    ],
    "company": [
        "div.job-details-jobs-unified-top-card__company-name a",
        "span.job-details-jobs-unified-top-card__company-name",
        "a.topcard__org-name-link",
        "span[data-test-id='job-poster-company-name']",
    ],
    "location": [
        "div.job-details-jobs-unified-top-card__primary-description-without-tagline span.tvm__text",
        "span.job-details-jobs-unified-top-card__bullet",
        "span[data-test-id='job-location']",
        "div.jobs-unified-top-card__primary-description > div > span",
    ],
    "employment_type": [
        "span.job-details-jobs-unified-top-card__job-insight-text",
        "li.job-details-jobs-unified-top-card__job-insight span",
        "span[data-test-id='job-employment-type']",
    ],
    "description": [
        "div.jobs-description__content",
        "div#job-details",
        "article.jobs-description__container",
        "div[class*='description']",
    ],
    "applicants": [
        "span.jobs-unified-top-card__applicant-count",
        "figcaption.jobs-unified-top-card__applicant-count",
    ],
    "posted_date": [
        "span.jobs-unified-top-card__posted-date",
        "time",
        "span[data-test-id='job-posted-date']",
    ],
    "seniority": [
        "span.job-details-jobs-unified-top-card__job-insight-text",
    ],
    "next_page": [
        "button[aria-label='View next page']",
        "button.jobs-search-pagination__button--next",
        "li.artdeco-pagination__indicator--number.selected + li button",
    ],
    "page_state": [
        "li.artdeco-pagination__indicator--number.selected",
    ],
}


def try_find(driver, selectors, parent=None, multiple=False, text_only=True):
    """Try multiple CSS selectors and return first successful match.

    Args:
        driver: Selenium WebDriver instance
        selectors: List of CSS selectors to try
        parent: Optional parent element to search within
        multiple: If True, return all matches as list; if False, return first match
        text_only: If True, return text content; if False, return element objects

    Returns:
        Extracted text/elements, or empty string/list if nothing found
    """
    ctx = parent if parent else driver
    for sel in selectors:
        try:
            if multiple:
                els = ctx.find_elements(By.CSS_SELECTOR, sel)
                if els:
                    return [e.text.strip() for e in els] if text_only else els
            else:
                el = ctx.find_element(By.CSS_SELECTOR, sel)
                if el:
                    return el.text.strip() if text_only else el
        except (NoSuchElementException, StaleElementReferenceException):
            continue
    return [] if multiple else ""


def build_driver(headless=False, user_data_dir=None):
    """Create and configure Chrome WebDriver."""
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")

    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--window-size=1440,900")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
    if user_data_dir:
        opts.add_argument(f"--user-data-dir={user_data_dir}")

    # Try cached chromedriver first, fallback to download if needed
    try:
        from pathlib import Path

        cache_path = Path.home() / ".wdm" / "drivers" / "chromedriver" / "win64"
        if cache_path.exists():
            # Get the latest cached version
            versions = sorted([d for d in cache_path.iterdir() if d.is_dir()], reverse=True)
            if versions:
                driver_path = versions[0] / "chromedriver-win32" / "chromedriver.exe"
                if driver_path.exists():
                    service = Service(str(driver_path))
                else:
                    service = Service(ChromeDriverManager().install())
            else:
                service = Service(ChromeDriverManager().install())
        else:
            service = Service(ChromeDriverManager().install())
    except (OSError, ValueError, FileNotFoundError):
        service = Service(ChromeDriverManager().install())

    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def wait_for_jobs_list(driver, timeout=15):
    """Wait until at least one job card is present."""
    wait = WebDriverWait(driver, timeout)
    for sel in SELECTORS["job_cards"]:
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
            return sel
        except TimeoutException:
            continue
    return None


def get_job_cards(driver, working_selector=None):
    """Return all job card elements on the current page."""
    selectors = (
        ([working_selector] + SELECTORS["job_cards"])
        if working_selector
        else SELECTORS["job_cards"]
    )
    for sel in selectors:
        cards = driver.find_elements(By.CSS_SELECTOR, sel)
        if cards:
            return cards, sel
    return [], None


def scroll_into_view_and_click(driver, element):
    """Scroll element into view and click it."""
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    time.sleep(0.3)
    try:
        element.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", element)


# ─────── Skill Extraction ─────────────────────────────────────────────────────


def _extract_skills(description: str):
    """Extract simple raw skill tokens from the description text.

    Args:
        description: Job description text

    Returns:
        List of detected skill keywords (lowercase)
    """
    if not description:
        return []

    patterns = [
        r"\bpython\b",
        r"\bjava\b",
        r"\bc\+\+\b",
        r"\bc#\b",
        r"\bjavascript\b",
        r"\btypescript\b",
        r"\breact\b",
        r"\bnode\.js\b",
        r"\bdjango\b",
        r"\bflask\b",
        r"\bsql\b",
        r"\bpostgresql\b",
        r"\bmysql\b",
        r"\baws\b",
        r"\bazure\b",
        r"\bdocker\b",
        r"\bkubernetes\b",
        r"\bgit\b",
    ]

    text = description.lower()
    found = []
    for pattern in patterns:
        if re.search(pattern, text):
            found.append(re.sub(r"\\b", "", pattern).replace("\\", ""))
    return found


# ─────── Role Filtering Keywords ──────────────────────────────────────────────

POSITIVE_ROLE_KEYWORDS = [
    "yazilim",
    "yazılım",
    "yazilim muhendisi",
    "yazılım mühendisi",
    "bilgisayar muhendisi",
    "bilgisayar mühendisi",
    "yapay zeka",
    "veri bilimi",
    "makine ogrenmesi",
    "makine öğrenmesi",
    "gelistirici",
    "geliştirici",
    "software",
    "developer",
    "engineer",
    "engineering",
    "backend",
    "frontend",
    "full stack",
    "fullstack",
    "data",
    "ai",
    "ml",
    "devops",
    "cloud",
    "cyber",
    "security",
    "robotics",
    "automation",
    "iot",
]

NEGATIVE_ROLE_KEYWORDS = [
    "satis",
    "satış",
    "satis temsil",
    "satış temsil",
    "musteri temsil",
    "müşteri temsil",
    "call center",
    "cagri merkezi",
    "çağrı merkezi",
    "pazarlama",
    "marketing",
    "client advisor",
    "tahsilat",
]

STRICT_TITLE_KEYWORDS = [
    "software",
    "yazılım",
    "yazilim",
    "developer",
    "backend",
    "frontend",
    "fullstack",
    "full stack",
    "qa",
    "test",
    "tester",
    "quality assurance",
    "analyst",
    "specialist",
    "system",
    "data",
    "devops",
    "cloud",
    "cyber",
    "security",
    "ai",
    "ml",
    "automation",
    "robotics",
    "iot",
    "it",
    "tech",
    "technology",
]

REQUIRED_SECTION_HEADERS = [
    "Aranan Nitelikler",
    "Gereksinimler",
    "Nitelikler",
    "Qualifications",
    "Requirements",
    "What we are looking for",
]


def _contains_any(text: str, keywords):
    """Check if text contains any of the keywords (case-insensitive)."""
    text = (text or "").lower()
    return any(keyword in text for keyword in keywords)


def is_relevant_job(title: str, description: str = ""):
    """Return (is_relevant, reason) for role-level filtering.

    Args:
        title: Job title
        description: Job description

    Returns:
        Tuple of (bool, str) - (is_relevant, reason_if_filtered)
    """
    title_l = (title or "").lower()
    text_l = f"{title_l} {(description or '').lower()}"

    if _contains_any(title_l, NEGATIVE_ROLE_KEYWORDS):
        return False, "negative_title_keyword"

    if not _contains_any(text_l, STRICT_TITLE_KEYWORDS):
        return False, "missing_strict_title_keyword"

    return True, "ok"


def _extract_card_job_url(card, driver):
    """Return the LinkedIn job-detail URL from the job card if available."""
    try:
        links = card.find_elements(By.CSS_SELECTOR, "a[href*='/jobs/view/']")
        for link in links:
            href = link.get_attribute("href") or ""
            if "/jobs/view/" in href:
                return href
    except Exception:
        pass

    try:
        link = card.find_element(By.CSS_SELECTOR, "a")
        href = link.get_attribute("href") or ""
        if "/jobs/view/" in href:
            return href
    except Exception:
        pass

    current = getattr(driver, "current_url", "") or ""
    if "/jobs/view/" in current:
        return current
    return ""


def _extract_sections(description: str, headers):
    """Extract text blocks that start with any given section header."""
    if not description:
        return []

    lines = [ln.strip() for ln in description.splitlines()]
    blocks = []
    current = []
    capturing = False

    stop_markers = [
        "sorumluluk",
        "iş tanımı",
        "is tanimi",
        "about",
        "benefit",
        "yan hak",
        "tercih edilen",
        "preferred",
    ]
    headers_l = [h.lower() for h in headers]

    for line in lines:
        if not line:
            if capturing and current:
                blocks.append("\n".join(current).strip())
                current = []
                capturing = False
            continue

        ll = line.lower().strip(" :")
        if any(h.lower() in ll for h in headers_l):
            if capturing and current:
                blocks.append("\n".join(current).strip())
                current = []
            capturing = True
            continue

        if capturing and any(marker in ll for marker in stop_markers):
            if current:
                blocks.append("\n".join(current).strip())
            current = []
            capturing = False
            continue

        if capturing:
            current.append(line)

    if capturing and current:
        blocks.append("\n".join(current).strip())

    return [b for b in blocks if b]


def _extract_required_skills(description: str):
    """Extract required skills using Turkish/English requirement sections and token patterns."""
    if not description:
        return []

    sections = _extract_sections(description, REQUIRED_SECTION_HEADERS)
    candidate_text = "\n".join(sections) if sections else description

    skills = []
    seen = set()

    # Capture line-level bullet-like requirement items.
    for line in candidate_text.splitlines():
        clean = re.sub(r"^[\s\-\*\u2022\d\.)\(]+", "", line).strip(" :;")
        if len(clean) < 2 or len(clean) > 120:
            continue
        if any(h.lower() == clean.lower() for h in REQUIRED_SECTION_HEADERS):
            continue
        if clean.lower() in seen:
            continue
        if re.search(r"[a-zA-ZçğıöşüÇĞİÖŞÜ]", clean):
            skills.append(clean)
            seen.add(clean.lower())

    # Add normalized technical terms from regex-based extraction.
    for token in _extract_skills(candidate_text):
        if token.lower() not in seen:
            skills.append(token)
            seen.add(token.lower())

    return skills



def scrape_job_detail(driver, default_keyword="Yazılım|Mühendislik"):
    """Extract all available info from the currently-open job detail panel."""
    time.sleep(random.uniform(1.0, 2.0))

    job = {}
    job["platform"] = "linkedin"
    job["title"] = try_find(driver, SELECTORS["job_title"])
    job["company"] = try_find(driver, SELECTORS["company"])

    locs = try_find(driver, SELECTORS["location"], multiple=True)
    job["location"] = " · ".join(locs) if locs else ""

    insights = try_find(driver, SELECTORS["employment_type"], multiple=True)
    job["employment_type"] = insights[0] if insights else ""

    desc_el = try_find(driver, SELECTORS["description"], text_only=False)
    description = desc_el.text.strip()[:5000] if desc_el else ""

    raw_skills = _extract_required_skills(description)

    job["url"] = driver.current_url
    job["scraped_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    job["full_description"] = description
    job["required_skills_raw"] = json.dumps(raw_skills, ensure_ascii=False)
    job["preferred_skills_raw"] = json.dumps([], ensure_ascii=False)
    job["experience_years"] = ""
    job["education_level"] = ""
    job["salary_range"] = ""
    job["department"] = ""
    job["search_keyword"] = default_keyword

    return job


def scrape_page(driver, working_sel, verbose=True, seen_urls: set | None = None):
    """Scrape all job cards on the current page."""
    if seen_urls is None:
        seen_urls = set()

    jobs = []
    filtered_out = 0
    dupes = 0
    cards, working_sel = get_job_cards(driver, working_sel)

    if not cards:
        print("  No job cards found on this page.")
        return jobs, working_sel

    print(f"  Found {len(cards)} job cards.")

    for i, card in enumerate(cards, 1):
        try:
            card_title = ""
            for title_sel in [
                "h3",
                "a.job-card-list__title",
                "a[data-control-name='job_card_title']",
                "a",
            ]:
                try:
                    card_title = card.find_element(By.CSS_SELECTOR, title_sel).text.strip()
                    if card_title:
                        break
                except NoSuchElementException:
                    continue

            # Dedup check before clicking
            card_url = _extract_card_job_url(card, driver)
            norm = _normalize_job_url(card_url)
            if norm and norm in seen_urls:
                dupes += 1
                if verbose:
                    print(f"  [{i}/{len(cards)}] SKIP (duplicate): {card_title[:55]}")
                continue

            if verbose:
                print(f"  [{i}/{len(cards)}] Clicking: {card_title[:60] or '(untitled)'}")

            scroll_into_view_and_click(driver, card)
            time.sleep(random.uniform(1.5, 2.5))

            job_data = scrape_job_detail(driver)
            if card_url:
                job_data["url"] = card_url

            if not job_data["title"] and card_title:
                job_data["title"] = card_title

            is_relevant, reason = is_relevant_job(
                job_data.get("title", ""),
                job_data.get("full_description", ""),
            )
            if not is_relevant:
                filtered_out += 1
                if verbose:
                    print(f"    Skipped by role filter: {reason}")
                continue

            final_url = job_data.get("url", "")
            final_norm = _normalize_job_url(final_url)
            if final_norm and final_norm in seen_urls:
                dupes += 1
                if verbose:
                    print(f"    SKIP (duplicate after detail): {job_data.get('title','')[:50]}")
                continue

            seen_urls.add(final_norm or card_title.lower())
            jobs.append(job_data)

        except (StaleElementReferenceException, Exception) as e:
            print(f"  ⚠ Error on card {i}: {e}")
            cards, working_sel = get_job_cards(driver, working_sel)
            continue

    if verbose:
        if filtered_out:
            print(f"  Filtered out {filtered_out} non-relevant jobs.")
        if dupes:
            print(f"  Skipped {dupes} duplicate jobs.")

    return jobs, working_sel


def go_to_next_page(driver):
    """Click the next-page button. Returns True on success."""
    for sel in SELECTORS["next_page"]:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, sel)
            if btn and btn.is_enabled() and btn.is_displayed():
                scroll_into_view_and_click(driver, btn)
                time.sleep(random.uniform(2.0, 3.5))
                return True
        except (NoSuchElementException, ElementClickInterceptedException):
            continue
    return False


def save_to_excel(jobs, output_path):
    """Save jobs in Kariyer.net-compatible schema."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    columns = [
        ("Platform", 14),
        ("Title", 35),
        ("Company", 28),
        ("Location", 24),
        ("URL", 55),
        ("Scraped At", 20),
        ("Full Description", 90),
        ("Required Skills", 35),
        ("Preferred Skills", 35),
        ("Experience Years", 16),
        ("Education Level", 20),
        ("Salary Range", 20),
        ("Employment Type", 20),
        ("Department", 20),
    ]

    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    alt_fill = PatternFill("solid", start_color="EBF3FB")
    normal_fill = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0A66C2", underline="single")
    wrap_align = Alignment(vertical="top", wrap_text=True)

    field_map = [
        "platform",
        "title",
        "company",
        "location",
        "url",
        "scraped_at",
        "full_description",
        "required_skills_raw",
        "preferred_skills_raw",
        "experience_years",
        "education_level",
        "salary_range",
        "employment_type",
        "department",
    ]

    for row_idx, job in enumerate(jobs, 2):
        fill = alt_fill if row_idx % 2 == 0 else normal_fill
        ws.row_dimensions[row_idx].height = 60

        for col_idx, field in enumerate(field_map, 1):
            value = job.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = wrap_align
            cell.border = thin_border
            cell.font = link_font if field == "url" else data_font

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "LinkedIn Scrape Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial", color="0A66C2")
    ws2["A3"] = "Total jobs scraped:"
    ws2["B3"] = len(jobs)
    ws2["A4"] = "Scrape date:"
    ws2["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2["A6"] = "Top Companies"
    ws2["A6"].font = Font(bold=True, name="Arial")

    companies = Counter(j.get("company", "Unknown") for j in jobs)
    for i, (company, count) in enumerate(companies.most_common(20), 7):
        ws2[f"A{i}"] = company
        ws2[f"B{i}"] = count

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12

    wb.save(output_path)
    print(f"\nSaved {len(jobs)} jobs to {output_path}")


def _handle_authwall(driver, url):
    """Prompt for manual login if LinkedIn redirects to authwall. Returns True on success."""
    if "authwall" not in driver.current_url and "login" not in driver.current_url:
        return True

    print("\nAuth wall or login detected on LinkedIn.")
    print("1) Complete login manually in the opened browser.")
    print("2) After landing on the jobs page, come back here and press Enter.")
    print("3) For future runs: pass --user-data-dir with your logged-in Chrome profile.")
    input("Press Enter after login to continue (or Ctrl+C to stop): ")

    for _ in range(6):
        driver.get(url)
        time.sleep(3)
        if "authwall" not in driver.current_url and "login" not in driver.current_url:
            if wait_for_jobs_list(driver, timeout=5):
                return True

    print("\nWARNING: Login was not confirmed. Try again with --user-data-dir.")
    return False


def main():
    parser = argparse.ArgumentParser(description="LinkedIn Job Scraper")
    parser.add_argument(
        "--url",
        default="",
        help="Single LinkedIn search URL (overrides --keywords if set)",
    )
    parser.add_argument(
        "--keywords",
        default="",
        help=(
            "Comma-separated search terms, e.g. 'Backend Developer,Data Scientist'. "
            "If omitted, keywords are derived from ESCO digital/green skill collections."
        ),
    )
    parser.add_argument("--pages", type=int, default=5, help="Max pages per keyword (default: 5)")
    parser.add_argument("--output", default="data/linkedin_jobs.xlsx", help="Output Excel file")
    parser.add_argument("--headless", action="store_true", help="Run Chrome in headless mode")
    parser.add_argument("--wait", type=int, default=20, help="Page load timeout (default: 20s)")
    parser.add_argument("--user-data-dir", help="Chrome user data dir for persisted login session")
    parser.add_argument(
        "--dedup-from",
        default="",
        help=(
            "Comma-separated Excel files to load existing job URLs from "
            "(e.g. 'outputs/skillab_tech_final.xlsx'). Duplicate jobs are skipped."
        ),
    )
    parser.add_argument(
        "--max-jobs",
        type=int,
        default=0,
        help="Stop after collecting this many jobs and save. 0 = unlimited.",
    )
    args = parser.parse_args()

    # Build keyword → URL list
    if args.url:
        search_targets = [("custom", args.url)]
    elif args.keywords:
        kws = [k.strip() for k in args.keywords.split(",") if k.strip()]
        search_targets = [(kw, build_keyword_url(kw)) for kw in kws]
    else:
        esco_keywords = load_esco_keywords()
        search_targets = [(kw, build_keyword_url(kw)) for kw in esco_keywords]

    # Load existing URLs for deduplication
    dedup_paths = [p.strip() for p in args.dedup_from.split(",") if p.strip()]
    seen_urls: set[str] = load_existing_urls(dedup_paths)
    print(f"  Dedup: {len(seen_urls)} existing job URLs loaded from {len(dedup_paths)} file(s).")

    print("=" * 60)
    print("  LinkedIn Job Scraper — Multi-Keyword Mode")
    print("=" * 60)
    print(f"  Keywords : {len(search_targets)}")
    print(f"  Pages/kw : {args.pages}")
    print(f"  Max jobs : {args.max_jobs if args.max_jobs else 'unlimited'}")
    print(f"  Output   : {args.output}")
    print(f"  Headless : {args.headless}")
    print("=" * 60)
    print()

    if not args.headless:
        print("TIP: Sign in manually if LinkedIn shows authwall.")
        print("     Use --user-data-dir to persist your session across runs.")
        print()

    driver = build_driver(headless=args.headless, user_data_dir=args.user_data_dir)
    all_jobs: list[dict] = []
    logged_in_once = False

    try:
        for kw_idx, (keyword, url) in enumerate(search_targets, 1):
            print(f"\n{'='*60}")
            print(f"  KEYWORD {kw_idx}/{len(search_targets)}: {keyword}")
            print(f"{'='*60}")

            driver.get(url)
            time.sleep(random.uniform(3, 5))

            if not logged_in_once:
                if not _handle_authwall(driver, url):
                    break
                logged_in_once = True
            elif "authwall" in driver.current_url or "login" in driver.current_url:
                print("  Session expired — re-navigating to URL...")
                driver.get(url)
                time.sleep(3)

            working_sel = None
            for page_num in range(1, args.pages + 1):
                print(f"\n  {'-'*46}")
                print(f"  PAGE {page_num}  [{keyword}]")
                print(f"  {'-'*46}")

                working_sel = wait_for_jobs_list(driver, timeout=args.wait)
                if not working_sel:
                    print("  Could not find job cards. Skipping to next keyword.")
                    break

                page_jobs, working_sel = scrape_page(
                    driver, working_sel, seen_urls=seen_urls
                )
                all_jobs.extend(page_jobs)
                print(
                    f"  Page {page_num}: +{len(page_jobs)} new jobs. "
                    f"Total so far: {len(all_jobs)}"
                )

                if args.max_jobs and len(all_jobs) >= args.max_jobs:
                    print(f"\n  Max jobs limit ({args.max_jobs}) reached. Stopping.")
                    raise KeyboardInterrupt

                if page_num < args.pages:
                    print("  Navigating to next page...")
                    if not go_to_next_page(driver):
                        print("  No more pages.")
                        break
                    time.sleep(random.uniform(2, 4))

            if args.max_jobs and len(all_jobs) >= args.max_jobs:
                break

            # Polite pause between keywords
            if kw_idx < len(search_targets):
                pause = random.uniform(4, 8)
                print(f"\n  Pausing {pause:.1f}s before next keyword...")
                time.sleep(pause)

    except KeyboardInterrupt:
        print("\nInterrupted by user.")
    finally:
        driver.quit()

    if all_jobs:
        save_to_excel(all_jobs, args.output)
        print(f"\nTotal unique jobs scraped: {len(all_jobs)}")
    else:
        print("\nNo jobs scraped. Check login status or URL.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
