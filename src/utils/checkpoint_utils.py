"""
Checkpoint and Excel Append utilities for resilient 985-job scraping.

Manages:
  - Checkpoint saving/loading (page + job index for resume capability)
  - Excel appending (openpyxl-based, immediate writes after each job)
  - Audio alerts (winsound beep on Cloudflare detection)
  - Progress logging to terminal

Author: Skillab Turkey Team
Project: EU Horizon Skill Intelligence Hub
"""

from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Dict, Optional, Tuple
import json
import time

try:
    import winsound
except ImportError:
    winsound = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    openpyxl = None

import pandas as pd


CHECKPOINT_FILE: Path = Path("data/checkpoints/scrape_checkpoint_985.json")
EXCEL_APPEND_PATH: Path = Path("data/TURKIYE_YAZILIM_985_FINAL.xlsx")
FULL_RECOVERY_CSV_PATH: Path = Path("data/TURKIYE_YAZILIM_FULL_985.csv")


def ensure_excel_file(output_path: Path = EXCEL_APPEND_PATH) -> None:
    """Create Excel file with headers if it doesn't exist."""
    if openpyxl is None:
        return
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    if output_path.exists():
        return
    
    # Create workbook with headers
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"
    
    headers = [
        "Platform",
        "Title",
        "Company",
        "Location",
        "URL",
        "Scraped At",
        "Full Description",
        "Required Skills",
        "Preferred Skills",
        "Experience Years",
        "Education Level",
        "Salary Range",
        "Employment Type",
        "Department",
        "Digital Concepts",
        "Green Concepts",
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 15
    ws.column_dimensions["N"].width = 15
    ws.column_dimensions["O"].width = 25
    ws.column_dimensions["P"].width = 25
    
    wb.save(output_path)


def append_job_to_excel(
    job_data: Dict,
    output_path: Path = EXCEL_APPEND_PATH,
) -> None:
    """Append a single job record to Excel file immediately."""
    if openpyxl is None:
        return
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    ensure_excel_file(output_path)
    
    try:
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Prepare row data
        row_data = [
            job_data.get("platform", ""),
            job_data.get("title", ""),
            job_data.get("company", ""),
            job_data.get("location", ""),
            job_data.get("url", ""),
            job_data.get("scraped_at", ""),
            job_data.get("full_description", ""),
            job_data.get("required_skills_raw", ""),
            job_data.get("preferred_skills_raw", ""),
            job_data.get("experience_years", ""),
            job_data.get("education_level", ""),
            job_data.get("salary_range", ""),
            job_data.get("employment_type", ""),
            job_data.get("department", ""),
            job_data.get("digital_concepts", ""),
            job_data.get("green_concepts", ""),
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        wb.save(output_path)
    except Exception as e:
        print(f"[EXCEL_APPEND] Error appending to {output_path}: {e}")


def append_job_to_recovery_csv(
    job_data: Dict,
    output_path: Path = FULL_RECOVERY_CSV_PATH,
) -> None:
    """Append a single job row to recovery CSV immediately."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    row = pd.DataFrame([job_data])
    if output_path.exists():
        row.to_csv(output_path, mode="a", header=False, index=False, encoding="utf-8-sig")
    else:
        row.to_csv(output_path, index=False, encoding="utf-8-sig")


def reset_recovery_csv(output_path: Path = FULL_RECOVERY_CSV_PATH) -> None:
    """Delete old recovery CSV to start a fresh recovery run."""
    if output_path.exists():
        output_path.unlink()


def save_checkpoint(
    keyword: str,
    current_page: int,
    current_job_index: int,
) -> None:
    """Save scraping checkpoint for resume capability."""
    CHECKPOINT_FILE.parent.mkdir(parents=True, exist_ok=True)
    
    checkpoint = {
        "keyword": keyword,
        "page": current_page,
        "job_index": current_job_index,
        "timestamp": time.time(),
    }
    
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(checkpoint, f, indent=2)


def load_checkpoint() -> Optional[Dict]:
    """Load previous checkpoint or None if starting fresh."""
    if not CHECKPOINT_FILE.exists():
        return None
    
    try:
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[CHECKPOINT] Error loading checkpoint: {e}")
        return None


def clear_checkpoint() -> None:
    """Clear checkpoint after successful completion."""
    if CHECKPOINT_FILE.exists():
        CHECKPOINT_FILE.unlink()


def log_progress(job_index: int, total_jobs: int, company: str = "", title: str = "") -> None:
    """Log progress to terminal in format: [İlerleme: N / 985] - [Company] taranıyor..."""
    display_title = title[:40] if title else "..."
    message = f"[İlerleme: {job_index} / {total_jobs}] - [{company}] ({display_title}) taranıyor..."
    print(message, flush=True)


def beep_alert(reason: str = "Cloudflare Detected") -> None:
    """Play system beep sound to alert user."""
    if winsound is None:
        print(f"[ALERT] {reason} - Audio not available on this system")
        return
    
    try:
        # 3 beeps, 500ms duration, 1000Hz frequency
        for _ in range(3):
            winsound.Beep(1000, 500)
            time.sleep(0.2)
        print(f"[ALERT] {reason} - Sound played")
    except Exception as e:
        print(f"[ALERT] {reason} - Error playing sound: {e}")


def detect_cloudflare_challenge(page_source: str) -> bool:
    """Detect Cloudflare challenge markers in HTML."""
    cloudflare_markers = [
        "cf_challenge",
        "cf-challenge",
        "ray id",
        "ray=",
        "cloudflare",
        "checking your browser",
        "challenge",
        "__cf_bm",
    ]
    
    source_lower = page_source.lower()
    return any(marker in source_lower for marker in cloudflare_markers)
