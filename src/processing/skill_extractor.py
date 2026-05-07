"""
Skill extraction pipeline for Skillab Turkiye.

Pipeline position:  Stage 2 — Text Processing & Skill Extraction
Input:              LinkedIn Excel export (Full Description column)
Output:             data/skill_inventory.xlsx  (two sheets)

  Sheet 1 – Skill Inventory : human-readable, bulleted skill lists per job.
  Sheet 2 – ESCO Export     : machine-readable JSON skill arrays, ready for
                              an NLP model to map each skill to an ESCO URI.

Performance: bulk cleaning operations run as vectorised pandas string
  operations to support 20 000+ row datasets. All regex patterns are
  pre-compiled at module load time to avoid repeated compilation overhead.

Author: Skillab Turkey Team
Project: EU Horizon Skill Intelligence Hub
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Pattern

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from src.utils.logger import get_logger

logger = get_logger(__name__)

INPUT_PATH = Path("data/skill_inventory.xlsx")
OUTPUT_PATH = Path("data/skill_inventory.xlsx")
ESCO_DEFAULT_PATH = Path("data/esco/skills_en.csv")


# Refined description range delimiters.
# Priority: if "İş İlanı Hakkında" is not found, fallback markers are used.
START_MARKER_PATTERNS: List[Pattern[str]] = [
    re.compile(r"iş\s+ilanı\s+hakkında", re.IGNORECASE),
    re.compile(r"aranan\s+nitelikler", re.IGNORECASE),
    re.compile(r"genel\s+nitelikler", re.IGNORECASE),
    re.compile(r"iş\s+tanımı", re.IGNORECASE),
    re.compile(r"qualifications", re.IGNORECASE),
    re.compile(r"requirements", re.IGNORECASE),
]

END_MARKER_PATTERNS: List[Pattern[str]] = [
    re.compile(r"aday\s+kriterleri", re.IGNORECASE),
    re.compile(r"şirket\s+hakkında", re.IGNORECASE),
    re.compile(r"site\s+kullanımı", re.IGNORECASE),
    re.compile(r"veri\s+politikamız", re.IGNORECASE),
    re.compile(r"copyright", re.IGNORECASE),
]

# Company culture / branding sections that should not be part of skill extraction.
CULTURAL_NOISE_PATTERNS: List[Pattern[str]] = [
    re.compile(r"vizyonumuz", re.IGNORECASE),
    re.compile(r"misyonumuz", re.IGNORECASE),
    re.compile(r"biz\s+kimiz", re.IGNORECASE),
    re.compile(r"değerli\s+bir\s+üyesi", re.IGNORECASE),
    re.compile(r"kurulduğumuz\s+günden\s+bu\s+yana", re.IGNORECASE),
    re.compile(r"(şirket|kurum|biz)\s+hakkında|hakkımızda", re.IGNORECASE),
]

NOISE_SENTENCE_PATTERN: Pattern[str] = re.compile(
    r"yayınlandı|gün\s+önce|başvuru|şirket\s+ilanları", re.IGNORECASE
)


SKILL_PATTERNS: Dict[str, Pattern[str]] = {
    "Python": re.compile(r"\bpython\b", re.IGNORECASE),
    "SQL": re.compile(r"\bsql\b|\btsql\b|\bpl/sql\b", re.IGNORECASE),
    "Java": re.compile(r"\bjava\b", re.IGNORECASE),
    ".NET": re.compile(r"\b\.net\b|\basp\.net\b|\bdotnet\b", re.IGNORECASE),
    "C#": re.compile(r"\bc#\b", re.IGNORECASE),
    "JavaScript": re.compile(r"\bjavascript\b|\bjs\b", re.IGNORECASE),
    "TypeScript": re.compile(r"\btypescript\b|\bts\b", re.IGNORECASE),
    "HTML": re.compile(r"\bhtml5?\b", re.IGNORECASE),
    "CSS": re.compile(r"\bcss3?\b", re.IGNORECASE),
    "React": re.compile(r"\breact\b", re.IGNORECASE),
    "Angular": re.compile(r"\bangular\b", re.IGNORECASE),
    "Vue": re.compile(r"\bvue\b", re.IGNORECASE),
    "Node.js": re.compile(r"\bnode\.?js\b", re.IGNORECASE),
    "Git": re.compile(r"\bgit\b|\bgithub\b|\bsvn\b|\btfs\b", re.IGNORECASE),
    "Docker": re.compile(r"\bdocker\b", re.IGNORECASE),
    "Kubernetes": re.compile(r"\bkubernetes\b|\bk8s\b", re.IGNORECASE),
    "REST API": re.compile(r"\brest\b|\brestful\b|\bweb api\b", re.IGNORECASE),
    "SOAP": re.compile(r"\bsoap\b", re.IGNORECASE),
    "Agile": re.compile(r"\bagile\b|\bçevik\b|\bscrum\b", re.IGNORECASE),
    "OOP": re.compile(r"\boop\b|\bnesne (?:tabanlı|yönelimli)\b", re.IGNORECASE),
    "Entity Framework": re.compile(r"\bentity framework\b|\bef core\b", re.IGNORECASE),
    "PostgreSQL": re.compile(r"\bpostgresql\b", re.IGNORECASE),
    "MySQL": re.compile(r"\bmysql\b", re.IGNORECASE),
    "NoSQL": re.compile(r"\bnosql\b", re.IGNORECASE),
    "Power BI": re.compile(r"\bpower\s?bi\b", re.IGNORECASE),
    "AWS": re.compile(r"\baws\b|\bamazon\s+web\s+services\b", re.IGNORECASE),
    "Azure": re.compile(r"\bazure\b", re.IGNORECASE),
    "GCP": re.compile(r"\bgcp\b|\bgoogle\s+cloud\s+platform\b", re.IGNORECASE),
    "Spring Boot": re.compile(r"\bspring\s*boot\b", re.IGNORECASE),
    "FastAPI": re.compile(r"\bfastapi\b", re.IGNORECASE),
    "Jenkins": re.compile(r"\bjenkins\b", re.IGNORECASE),
    "CI/CD": re.compile(r"\bci\s*/\s*cd\b|\bcicd\b|\bcontinuous\s+integration\b", re.IGNORECASE),
    "Scrum": re.compile(r"\bscrum\b", re.IGNORECASE),
    "Kanban": re.compile(r"\bkanban\b", re.IGNORECASE),
    "Jira": re.compile(r"\bjira\b", re.IGNORECASE),
    "Bitbucket": re.compile(r"\bbitbucket\b", re.IGNORECASE),
}

SOFT_SKILL_PATTERNS: Dict[str, Pattern[str]] = {
    "Iletisim": re.compile(r"\biletişim\b|\bcommunication\b", re.IGNORECASE),
    "Takim Calismasi": re.compile(
        r"\btakım\s+çalışması\b|\bteamwork\b|\bteam\s+player\b", re.IGNORECASE
    ),
    "Liderlik": re.compile(r"\bliderlik\b|\bleadership\b", re.IGNORECASE),
    "Problem Cozme": re.compile(r"\bproblem\s+çözme\b|\bproblem\s+solving\b", re.IGNORECASE),
    "Yaraticilik": re.compile(r"\byaratıcılık\b|\bcreativity\b", re.IGNORECASE),
    "Analitik Dusunme": re.compile(
        r"\banalitik\s+düşünme\b|\banalytical\s+thinking\b", re.IGNORECASE
    ),
    "Zaman Yonetimi": re.compile(r"\bzaman\s+yönetimi\b|\btime\s+management\b", re.IGNORECASE),
}

COMPETENCY_CLUSTER_RULES: Dict[str, List[str]] = {
    "Yazilim Gelistirme": ["Python", "Java", "JavaScript", "TypeScript", ".NET", "C#"],
    "Web Teknolojileri": ["HTML", "CSS", "React", "Angular", "Vue", "Node.js", "REST API"],
    "Veri ve Veritabani": ["SQL", "PostgreSQL", "MySQL", "NoSQL", "Power BI"],
    "Bulut ve DevOps": ["AWS", "Azure", "GCP", "Docker", "Kubernetes", "CI/CD", "Jenkins"],
    "Surec ve Cevik": ["Agile", "Scrum", "Kanban", "Jira"],
}

TARGET_SKILL_KEYWORDS: List[str] = [
    "python",
    "javascript",
    "html",
    "css",
    "java",
    "sql",
    "api",
    "git",
    "github",
    "agile",
    "scrum",
    "bootstrap",
]

# ---------------------------------------------------------------------------
# Pre-compiled bulk-cleaning patterns (vectorised, column-level application)
# ---------------------------------------------------------------------------

_UI_TOKENS_REGEX: re.Pattern = re.compile(
    r"\b(keyboard_arrow_down|navigate_next|navigate_before|"
    r"thumb_up|thumb_down|chevron_right)\b",
    re.IGNORECASE,
)
_PHONE_REGEX: re.Pattern = re.compile(r"\b\d{3,4}[\s\-]?\d{2,4}[\s\-]?\d{2,4}\b")
_NBSP_REGEX: re.Pattern = re.compile(r"\u00a0+")
_WHITESPACE_REGEX: re.Pattern = re.compile(r"\s{2,}")
_PUNCTUATION_REGEX: re.Pattern = re.compile(r"[^\w\s]+", re.UNICODE)


def _normalize_whitespace(text: str) -> str:
    """Collapse all whitespace to single spaces."""
    return re.sub(r"\s+", " ", text).strip()


def preprocess_description_text(text: str) -> str:
    """
    Normalize text before analysis.

    Steps:
      1. Lowercase
      2. Remove punctuation
      3. Collapse excessive whitespace
    """
    if not isinstance(text, str):
        return ""
    normalized = text.lower()
    normalized = _PUNCTUATION_REGEX.sub(" ", normalized)
    return _normalize_whitespace(normalized)


def _slice_content_window(text: str) -> str:
    """
    Keep only text between start and end markers.

    Returns an empty string when a start marker cannot be found.
    """
    start_match = None
    for pattern in START_MARKER_PATTERNS:
        candidate = pattern.search(text)
        if candidate:
            start_match = candidate
            break

    if not start_match:
        return ""

    slice_start = start_match.start()
    sliced = text[slice_start:]

    earliest_end = len(sliced)
    for pattern in END_MARKER_PATTERNS:
        candidate = pattern.search(sliced, start_match.end() - slice_start)
        if candidate:
            earliest_end = min(earliest_end, candidate.start())

    return sliced[:earliest_end]


def _remove_cultural_noise(text: str) -> str:
    """Remove company-culture sentences from sliced text."""
    if not text:
        return ""

    # Remove section heading so the keyword "hakkında" does not drop useful content.
    text = re.sub(r"\biş\s+ilanı\s+hakkında\b", " ", text, flags=re.IGNORECASE)

    chunks = re.split(r"(?<=[.!?])\s+", text)
    kept_chunks: List[str] = []

    for chunk in chunks:
        if not chunk.strip():
            continue
        if any(pattern.search(chunk) for pattern in CULTURAL_NOISE_PATTERNS):
            continue
        kept_chunks.append(chunk)

    return " ".join(kept_chunks)


def _remove_noise_sentences(text: str) -> str:
    """Remove operational/listing noise fragments from text."""
    if not text:
        return ""

    chunks = re.split(r"(?<=[.!?])\s+|\s*,\s*|\s*;\s*", text)
    kept_chunks: List[str] = []

    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue
        if NOISE_SENTENCE_PATTERN.search(chunk):
            continue
        kept_chunks.append(chunk)

    return ". ".join(kept_chunks)


def _clean_window(description: str) -> str:
    """
    Per-row cleaning: extract the relevant content window then strip
    cultural-branding and operational listing noise sentences.

    Bulk regex substitutions (portal labels, UI tokens, phone numbers,
    whitespace) are applied at the DataFrame level for large-dataset
    performance — see ``build_processed_dataset``.
    """
    if not isinstance(description, str):
        return ""
    text = _slice_content_window(description)
    text = _remove_cultural_noise(text)
    text = _remove_noise_sentences(text)
    return text


def _extract_skills_vectorized(descriptions: pd.Series) -> pd.Series:
    """
    Vectorised skill extraction across the entire column.

    For each skill, a single ``str.contains()`` call evaluates every row at
    once in C-speed pandas code. The boolean results are assembled into a
    DataFrame, then converted to sorted skill lists per row via one
    ``apply()`` on the small boolean frame — not on raw text.

    Complexity: O(skills) vectorised regex passes — scales to 20 000+ rows
    with near-constant overhead per additional skill added to the taxonomy.
    """
    flag_df = pd.DataFrame(
        {
            skill: descriptions.str.contains(
                pattern.pattern, flags=re.IGNORECASE, na=False, regex=True
            )
            for skill, pattern in SKILL_PATTERNS.items()
        },
        index=descriptions.index,
    )
    return flag_df.apply(lambda row: sorted(row.index[row].tolist()), axis=1)


def _extract_pattern_matches_vectorized(
    descriptions: pd.Series,
    pattern_map: Dict[str, Pattern[str]],
) -> pd.Series:
    """Return sorted pattern-key matches per row for the given pattern map."""
    flag_df = pd.DataFrame(
        {
            label: descriptions.str.contains(
                pattern.pattern, flags=re.IGNORECASE, na=False, regex=True
            )
            for label, pattern in pattern_map.items()
        },
        index=descriptions.index,
    )
    return flag_df.apply(lambda row: sorted(row.index[row].tolist()), axis=1)


def _build_competency_clusters(skills: List[str]) -> List[str]:
    """Map extracted hard skills to competency clusters."""
    detected_clusters: List[str] = []
    skill_set = set(skills)
    for cluster_name, cluster_skills in COMPETENCY_CLUSTER_RULES.items():
        if any(skill in skill_set for skill in cluster_skills):
            detected_clusters.append(cluster_name)
    return sorted(detected_clusters)


def _skills_to_bullets(skills: List[str]) -> str:
    """Render skill list as a multiline bullet string for Excel cells."""
    if not skills:
        return "-"
    return "\n".join(f"- {skill}" for skill in skills)


def build_skill_frequency_table(
    descriptions: pd.Series,
    keywords: List[str] = TARGET_SKILL_KEYWORDS,
) -> pd.DataFrame:
    """Build keyword frequency table from preprocessed descriptions."""
    if descriptions.empty:
        return pd.DataFrame(columns=["Yetenek", "Frekans"])

    frequencies = {
        keyword: int(
            descriptions.str.count(rf"\b{re.escape(keyword)}\b", flags=re.IGNORECASE).sum()
        )
        for keyword in keywords
    }
    frequency_df = pd.DataFrame(
        {
            "Yetenek": list(frequencies.keys()),
            "Frekans": list(frequencies.values()),
        }
    )
    return frequency_df.sort_values(by=["Frekans", "Yetenek"], ascending=[False, True]).reset_index(
        drop=True
    )


def build_processed_dataset(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a clean, ESCO-ready skill inventory from raw scraped data.

    Cleaning pipeline:
        1. Per-row   — content window extraction + cultural/listing noise removal.
        2. Vectorised — noise phrase stripping, UI tokens, whitespace normalisation.
        3. Vectorised — skill detection via one boolean mask per skill.

    ESCO readiness:
        ``skills_esco_ready`` stores each row's skills as a JSON array string.
        The Stage 3 NLP mapper ingests it with ``json.loads()`` and resolves
        each item to an ESCO skill URI without additional parsing.
    """
    if "full_description" not in df.columns:
        raise ValueError("Input CSV must contain a 'full_description' column.")

    work = df[
        [
            c
            for c in [
                "platform",
                "title",
                "company",
                "location",
                "url",
                "scraped_at",
                "education_level",
                "salary_range",
                "employment_type",
                "department",
                "full_description",
            ]
            if c in df.columns
        ]
    ].copy()
    work["full_description"] = work["full_description"].fillna("").astype(str)

    logger.info("cleaning_started", total_rows=len(work))

    # ── Step 1: per-row window + cultural/listing noise removal ──────────────
    work["cleaned_description"] = work["full_description"].map(_clean_window)

    # ── Step 2: vectorised bulk cleaning (fast on large datasets) ────────────
    col = work["cleaned_description"]
    col = col.str.replace(_UI_TOKENS_REGEX, " ", regex=True)
    col = col.str.replace(_PHONE_REGEX, " ", regex=True)
    col = col.str.replace(_NBSP_REGEX, " ", regex=True)
    col = col.str.replace(_WHITESPACE_REGEX, " ", regex=True).str.strip()
    work["cleaned_description"] = col

    # ── Step 3: preprocessing for consistent keyword analysis ────────────────
    work["normalized_description"] = work["cleaned_description"].map(preprocess_description_text)

    # ── Step 4: discard rows with no usable description ──────────────────────
    before = len(work)
    work = work[work["normalized_description"].str.len() > 20].copy()
    logger.info(
        "window_filter_applied",
        kept=len(work),
        dropped=before - len(work),
    )

    # ── Step 5: vectorised skill extraction ───────────────────────────────────
    work["extracted_skills"] = _extract_skills_vectorized(work["normalized_description"])
    work["soft_skills"] = _extract_pattern_matches_vectorized(
        work["normalized_description"], SOFT_SKILL_PATTERNS
    )
    work["skill_count"] = work["extracted_skills"].map(len)
    work["soft_skill_count"] = work["soft_skills"].map(len)

    # ── Step 6: discard non-technical rows (zero extractable skills) ──────────
    before = len(work)
    work = work[work["skill_count"] > 0].copy()
    logger.info(
        "skill_filter_applied",
        technical_rows=len(work),
        dropped=before - len(work),
    )

    # ── Step 7: build display and ESCO-ready representations ─────────────────
    work["skills_display"] = work["extracted_skills"].map(
        lambda skills: "\n".join(f"\u2022 {s}" for s in skills)
    )
    work["skills_esco_ready"] = work["extracted_skills"].map(
        lambda skills: json.dumps(skills, ensure_ascii=False)
    )
    work["soft_skills_display"] = work["soft_skills"].map(_skills_to_bullets)
    work["competency_clusters"] = work["extracted_skills"].map(_build_competency_clusters)
    work["competency_clusters_display"] = work["competency_clusters"].map(_skills_to_bullets)

    return work.reset_index(drop=True)


# ============================================================================
# Excel output helpers
# ============================================================================


def _style_header_row(ws) -> None:
    """Apply Skillab brand styling to the header row of a worksheet."""
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    centre = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = centre


def _format_inventory_sheet(ws) -> None:
    """Format the original postings sheet."""
    _style_header_row(ws)
    ws.column_dimensions["A"].width = 13  # Platform
    ws.column_dimensions["B"].width = 36  # Job Title
    ws.column_dimensions["C"].width = 30  # Company
    ws.column_dimensions["D"].width = 20  # Location
    ws.column_dimensions["E"].width = 11  # Hard Skill Count
    ws.column_dimensions["F"].width = 11  # Soft Skill Count
    ws.column_dimensions["G"].width = 38  # Hard Skills
    ws.column_dimensions["H"].width = 32  # Soft Skills
    ws.column_dimensions["I"].width = 34  # Competency Clusters
    ws.column_dimensions["J"].width = 18  # Scraped At
    ws.column_dimensions["K"].width = 35  # URL
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_top


def _format_analysis_sheet(ws) -> None:
    """Format the skill-frequency summary sheet."""
    _style_header_row(ws)
    ws.column_dimensions["A"].width = 30  # Skill
    ws.column_dimensions["B"].width = 18  # Frequency
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_top


def _write_excel(
    processed_df: pd.DataFrame,
    frequency_df: pd.DataFrame,
    path: Path,
) -> None:
    """
    Write the two-sheet Excel workbook.

    Sheet 1 – Tum_Ilanlar     : readable, entity-enriched postings table.
    Sheet 2 – Yetenek_Analizi : skill keyword frequencies.
    """
    readable_postings = processed_df[
        [
            c
            for c in [
                "platform",
                "title",
                "company",
                "location",
                "skill_count",
                "soft_skill_count",
                "skills_display",
                "soft_skills_display",
                "competency_clusters_display",
                "scraped_at",
                "url",
            ]
            if c in processed_df.columns
        ]
    ].rename(
        columns={
            "platform": "Platform",
            "title": "Pozisyon",
            "company": "Sirket",
            "location": "Lokasyon",
            "skill_count": "Hard_Skill_Sayisi",
            "soft_skill_count": "Soft_Skill_Sayisi",
            "skills_display": "Hard_Skills",
            "soft_skills_display": "Soft_Skills",
            "competency_clusters_display": "Yetkinlik_Kumeleri",
            "scraped_at": "Tarih",
            "url": "URL",
        }
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        readable_postings.to_excel(writer, index=False, sheet_name="Tum_Ilanlar")
        frequency_df.to_excel(writer, index=False, sheet_name="Yetenek_Analizi")
        _format_inventory_sheet(writer.book["Tum_Ilanlar"])
        _format_analysis_sheet(writer.book["Yetenek_Analizi"])


def _find_description_column(columns: List[str]) -> str:
    """Resolve the best-fit description column from an Excel dataset."""
    canonical = {str(col).strip().lower(): col for col in columns}
    candidates = [
        "description",
        "full_description",
        "full description",
        "job_description",
        "job description",
    ]
    for key in candidates:
        if key in canonical:
            return canonical[key]
    raise ValueError(
        "Input Excel must contain one of: description, full_description, full description, "
        "job_description, job description."
    )


def _build_esco_phrase_index(esco_csv_path: Path) -> tuple[Dict[str, Dict[str, str]], int]:
    """
    Build a normalized phrase index from ESCO CSV for digital/green skills only.

    Returns:
        - phrase_index: normalized phrase -> metadata dict
        - max_ngram_len: maximum phrase token length
    """
    esco_df = pd.read_csv(esco_csv_path)
    required_cols = {"conceptUri", "preferredLabel", "altLabels", "pillar"}
    missing = required_cols - set(esco_df.columns)
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"ESCO CSV is missing required columns: {missing_list}")

    phrase_index: Dict[str, Dict[str, str]] = {}
    max_ngram_len = 1

    for _, row in esco_df.iterrows():
        pillar = str(row.get("pillar", "")).strip().lower()
        if pillar not in {"digital", "green"}:
            continue

        concept_uri = str(row.get("conceptUri", "")).strip()
        preferred_label = str(row.get("preferredLabel", "")).strip()
        if not concept_uri or not preferred_label:
            continue

        terms = [preferred_label]
        alt_labels = str(row.get("altLabels", "") or "")
        if alt_labels and alt_labels.lower() != "nan":
            terms.extend([item.strip() for item in alt_labels.split("|") if item.strip()])

        for term in terms:
            normalized_term = preprocess_description_text(term)
            if not normalized_term:
                continue

            phrase_index.setdefault(
                normalized_term,
                {
                    "esco_preferred_label": preferred_label,
                    "esco_uri": concept_uri,
                    "label": pillar,
                },
            )
            max_ngram_len = max(max_ngram_len, len(normalized_term.split()))

    if not phrase_index:
        raise ValueError("No digital/green ESCO terms found in the provided taxonomy CSV.")

    return phrase_index, max_ngram_len


def _extract_esco_matches_from_description(
    normalized_description: str,
    phrase_index: Dict[str, Dict[str, str]],
    max_ngram_len: int,
) -> List[Dict[str, object]]:
    """Extract ESCO digital/green skill phrases using longest-first n-gram matching."""
    if not normalized_description:
        return []

    tokens = normalized_description.split()
    if not tokens:
        return []

    used = [False] * len(tokens)
    matches: List[Dict[str, object]] = []
    seen_uris: set[str] = set()

    n_max = min(max_ngram_len, len(tokens))
    for n in range(n_max, 0, -1):
        for i in range(0, len(tokens) - n + 1):
            if any(used[i : i + n]):
                continue

            phrase = " ".join(tokens[i : i + n])
            meta = phrase_index.get(phrase)
            if not meta:
                continue

            uri = meta["esco_uri"]
            if uri in seen_uris:
                continue

            for j in range(i, i + n):
                used[j] = True

            seen_uris.add(uri)
            matches.append(
                {
                    "skill": phrase,
                    "label": meta["label"],
                    "esco_preferred_label": meta["esco_preferred_label"],
                    "esco_uri": uri,
                    "match_method": "exact_phrase",
                    "match_score": 1.0,
                }
            )

    return matches


def run_esco_excel_labeling(
    input_excel_path: Path,
    output_excel_path: Path,
    esco_csv_path: Path = ESCO_DEFAULT_PATH,
    sheet_name: int | str = 0,
) -> pd.DataFrame:
    """
    Process the description column in an Excel file and label each extracted
    skill as digital/green using ESCO taxonomy.

    Writes a new Excel workbook with:
      1. `ESCO_Skill_Labels`     - row-level skill labels.
      2. `ESCO_Skill_Summary`    - grouped label frequencies.

    Args:
        input_excel_path: Source Excel file containing description text.
        output_excel_path: Destination Excel path for labeled result.
        esco_csv_path: ESCO taxonomy CSV path (skills_en.csv format).
        sheet_name: Sheet index or name to read from input workbook.

    Returns:
        Row-level labeled skill DataFrame.
    """
    logger.info(
        "esco_excel_labeling_started",
        input=str(input_excel_path),
        output=str(output_excel_path),
        esco_csv=str(esco_csv_path),
    )

    if not input_excel_path.exists():
        raise FileNotFoundError(f"Input Excel not found: {input_excel_path}")
    if not esco_csv_path.exists():
        raise FileNotFoundError(
            f"ESCO taxonomy CSV not found: {esco_csv_path}. "
            "Download ESCO skills CSV and place it at this path."
        )

    source_df = pd.read_excel(input_excel_path, sheet_name=sheet_name)
    if source_df.empty:
        raise ValueError("Input Excel sheet is empty.")

    desc_col = _find_description_column(list(source_df.columns))
    working_df = source_df.copy()
    working_df[desc_col] = working_df[desc_col].fillna("").astype(str)

    phrase_index, max_ngram_len = _build_esco_phrase_index(esco_csv_path)
    working_df["normalized_description"] = working_df[desc_col].map(preprocess_description_text)

    labeled_rows: List[Dict[str, object]] = []
    for idx, row in working_df.iterrows():
        description_text = row[desc_col]
        skill_matches = _extract_esco_matches_from_description(
            row["normalized_description"], phrase_index, max_ngram_len
        )
        for mapped_payload in skill_matches:
            labeled_rows.append(
                {
                    "source_row": int(idx) + 1,
                    "description": description_text,
                    **mapped_payload,
                }
            )

    labeled_df = pd.DataFrame(
        labeled_rows,
        columns=[
            "source_row",
            "description",
            "skill",
            "label",
            "esco_preferred_label",
            "esco_uri",
            "match_method",
            "match_score",
        ],
    )

    summary_df = (
        labeled_df.groupby(["skill", "label"], as_index=False)
        .size()
        .rename(columns={"size": "count"})
        .sort_values(by=["count", "skill"], ascending=[False, True])
    )

    output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        labeled_df.to_excel(writer, index=False, sheet_name="ESCO_Skill_Labels")
        summary_df.to_excel(writer, index=False, sheet_name="ESCO_Skill_Summary")
        _style_header_row(writer.book["ESCO_Skill_Labels"])
        _style_header_row(writer.book["ESCO_Skill_Summary"])

    logger.info(
        "esco_excel_labeling_completed",
        output=str(output_excel_path),
        rows=len(labeled_df),
        unique_skills=labeled_df["skill"].nunique() if not labeled_df.empty else 0,
    )
    return labeled_df


# ============================================================================
# Public pipeline entry point
# ============================================================================


def run_extraction(
    input_path: Path = INPUT_PATH,
    output_path: Path = OUTPUT_PATH,
) -> pd.DataFrame:
    """
    Run the full Stage 2 extraction pipeline and persist output to Excel.

    Args:
        input_path:  Path to input CSV with a full_description column.
        output_path: Destination path for the Excel workbook.

    Returns:
        Processed DataFrame for optional downstream use.

    Raises:
        FileNotFoundError: If ``input_path`` does not exist.
        RuntimeError:      If processing or Excel writing fails.
    """
    logger.info("extraction_started", input=str(input_path))

    if not input_path.exists():
        raise FileNotFoundError(f"Stage 1 output not found: {input_path}")

    try:
        raw_df = pd.read_csv(input_path)
    except Exception as exc:
        logger.error("csv_read_failed", path=str(input_path), error=str(exc))
        raise RuntimeError(f"Failed to read {input_path}: {exc}") from exc

    try:
        processed_df = build_processed_dataset(raw_df)
        skill_frequency_df = build_skill_frequency_table(processed_df["normalized_description"])
    except Exception as exc:
        logger.error("processing_failed", error=str(exc), exc_info=True)
        raise RuntimeError(f"Dataset processing failed: {exc}") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        _write_excel(processed_df, skill_frequency_df, output_path)
    except Exception as exc:
        logger.error("excel_write_failed", path=str(output_path), error=str(exc))
        raise RuntimeError(f"Excel export failed: {exc}") from exc

    logger.info(
        "extraction_completed",
        output=str(output_path),
        technical_jobs=len(processed_df),
        analyzed_skills=len(skill_frequency_df),
    )
    return processed_df


def main() -> None:
    """CLI entry point for CSV extraction or ESCO-labeled Excel processing."""
    parser = argparse.ArgumentParser(description="Skill extraction and ESCO labeling")
    parser.add_argument(
        "--excel-input",
        type=Path,
        help="Input Excel path containing description/full_description column.",
    )
    parser.add_argument(
        "--excel-output",
        type=Path,
        default=Path("data/esco_labeled_skills.xlsx"),
        help="Output Excel path for ESCO digital/green labels.",
    )
    parser.add_argument(
        "--esco-csv",
        type=Path,
        default=ESCO_DEFAULT_PATH,
        help="ESCO taxonomy CSV path (skills_en.csv).",
    )
    parser.add_argument(
        "--sheet",
        default=0,
        help="Excel sheet name or index for --excel-input mode.",
    )
    args = parser.parse_args()

    try:
        if args.excel_input:
            run_esco_excel_labeling(
                input_excel_path=args.excel_input,
                output_excel_path=args.excel_output,
                esco_csv_path=args.esco_csv,
                sheet_name=args.sheet,
            )
        else:
            run_extraction()
    except FileNotFoundError as exc:
        logger.error("input_not_found", detail=str(exc))
        raise SystemExit(1) from exc
    except RuntimeError as exc:
        logger.error("extraction_failed", detail=str(exc))
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
